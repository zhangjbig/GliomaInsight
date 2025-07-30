from PyQt5 import uic
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QDialog, QGraphicsDropShadowEffect, QFileDialog, QTableWidgetItem

import nibabel as nib
from pathlib import Path

from openpyxl.reader.excel import load_workbook

from RadiomicsPage import RadiomicsFunc
from VTK_new import SimpleView
from DlpFeatures import DLPExtractor
from lib.RadioMLPage import RadioML
from Script.Expression_Prediction import RunPredict

import matplotlib
matplotlib.use("Qt5Agg")  # 声明使用QT5
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

class MyFigure(FigureCanvas):
    def __init__(self, width=5, height=4, dpi=100):
        # 创建一个创建Figure
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        # 在父类中激活Figure窗口
        super(MyFigure, self).__init__(self.fig)  # 此句必不可少，否则不能显示图形
        # 创建一个子图，用于绘制图形用，111表示子图编号，如matlab的subplot(1,1,1)
        #self.axes = self.fig.add_subplot(111)

class Ui_MainDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.dia = uic.loadUi("New_UI_.ui", self)
        self.save_path = "./"
        self.initDia()

    def initDia(self):
        self.dia.textBrowser_2.append(
            "Glioma-GenomeDynamics is a free open source software platform for medical image progressing and 3D visualization of image data."
            "This mudule contains some basic methods to visualize and give anlaysis of medical image computing data sets.\n")
        self.dia.textBrowser_3.append(
            "You can perform basic traditional radiomics analysis by doing this:\n"
                "1.Click 'Input Label' and select your ROI\n"
                "2.Click 'Input Image' and select your brain image\n"
                "3.Choose Pyradiomics and select the settings you need. If you're not sure about that, commonSettings and choosing your output table path will be fine.\n"
                "4.You can name the output csv.\n"
                "5.Click the 'Apply' to see the visual analysis.If you're satisfied with that, click 'Save Data' to save it.")

        # 设置页面样式
        self.dia.setWindowFlag(Qt.FramelessWindowHint)  # 将界面设置为无框
        self.dia.setAttribute(Qt.WA_TranslucentBackground)  # 将界面属性设置为半透明
        self.dia.shadow = QGraphicsDropShadowEffect()  # 设定一个阴影,半径为10,颜色为#444444,定位为0,0
        self.dia.shadow.setBlurRadius(10)
        self.dia.shadow.setColor(QColor("#444444"))
        self.dia.shadow.setOffset(0, 0)
        self.dia.frame.setGraphicsEffect(self.dia.shadow)  # 为frame设定阴影效果
        self.dia.progressBar = self.progressBar
        self.progressBar.setMinimum(0)# 将进度条最小值设为0

        self.pushButton_close.clicked.connect(self.quit_button)
        self.dia.listWidget.itemClicked.connect(self.ChangePage)
        self.dia.listWidget_2.itemClicked.connect(self.ChangePage)

        #self.stackedWidget_3.setCurrentIndex(0)

        self.Actions()

    def Actions(self):
        # 初始化属性
        self.ori_path = ''
        self.ori_path_2 = ''
        self.ori_path_3 = ''
        self.ori_path_4 = ''
        self.save_path = ''
        self.RFs_file = "Script/Express_Visualize/test_example_data.csv"
        self.output_gene_path = ''
        self.output_gene_name = 'Gene_Expression_Result'
        self.image_files = {}
        self.feaNum = '20'
        # 上传图片
        self.data2 =''
        self.dia.pushButton.clicked.connect(self.bindButton)
        self.dia.pushButton_ori.clicked.connect(lambda: self.oriImg())  # 上传图像

        self.dia.pushButton_18.clicked.connect(lambda : self.stackedWidget_3.setCurrentIndex(0))
        self.dia.pushButton_19.clicked.connect(lambda : self.stackedWidget_3.setCurrentIndex(1))
        self.dia.pushButton_20.clicked.connect(lambda : self.stackedWidget_3.setCurrentIndex(2))
        self.dia.pushButton_21.clicked.connect(lambda : self.stackedWidget_3.setCurrentIndex(3))

        self.RadiomicsPart()
        self.DLPart()
        self.RadioML()
        self.PredictionPart()

    def getImage(self,path):
        try:
            data_nii = nib.load(path)
        except FileNotFoundError:
            pass
        else:
            data = data_nii.get_fdata()

        return data

    # 获取并展示全图
    def oriImg(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ExistingFiles
        files, _ = QFileDialog.getOpenFileNames(self, "QFileDialog.getOpenFileNames()", "./", "nii(*.nii.gz;*.nii;*.nrrd)", options=options)
        try:
            i = 0
            #for file_name in files:
            self.ori_path = files[0]  # 提取文件路径
            self.ori_path_2 = files[1]
            self.ori_path_3 = files[2]
            self.ori_path_4 = files[3]
            # 定义用于多模态分析的图像文件路径
            self.image_files = {
                'T1': self.ori_path,
                'T2': self.ori_path_2,
                'flair': self.ori_path_3,
                't1c': self.ori_path_4,
            }
            # 读取图像
            self.images = {}
            for modality, file_path in self.image_files.items():
                if file_path:
                    self.images[modality] = file_path
                    print(modality + ':' + file_path + '\n')
            print("images done")
        except nib.filebasedimages.ImageFileError | IndexError | RuntimeError | TypeError:
            pass
        else:
            self.Button_Slice()
            try:
                SimpleView.VisNii(self, self.nii_path, self.ori_path)
                # SimpleView.VisNii(self, self.nii_path, self.ori_path_2)
                # SimpleView.VisNii(self, self.nii_path, self.ori_path_3)
                # SimpleView.VisNii(self, self.nii_path, self.ori_path_4)
            except AttributeError:
                pass

    # 展示切片
    def showimage(self):
        self.data = self.getImage()

    #button中的切片
    def Button_Slice(self):
        def figer(place):  # place为继承容器
            # 定义MyFigure类的一个实例
            F = MyFigure(width=5, height=4, dpi=100)
            # 在GUI的groupBox中创建一个布局，用于添加MyFigure类的实例（即图形）后其他部件。
            place.addWidget(F, 0, 1)
            return F

        def upDateCavana(F, data, ori_slicer):
            width, height = self.graphicsView_2.width(), self.graphicsView_2.height()
            fig = F.figure
            fig.clear()
            F.resize(width, height)
            ax = fig.add_axes([0, 0, 1, 1])  # 3是宽度，4是高度
            ax.imshow(data[:, :, ori_slicer], cmap='gray')
            fig.canvas.draw()  # draw可以用于画布更新

        self.F = figer(self.gridLayout_001)
        self.F2 = figer(self.gridLayout_002)
        self.F3 = figer(self.gridLayout_003)
        self.F4 = figer(self.gridLayout_004)

        data1 = self.getImage(self.ori_path)
        ori_slicer1 = (int)(data1.shape[2] / 2)
        data2 = self.getImage(self.ori_path_2)
        ori_slicer2 = (int)(data2.shape[2] / 2)
        data3 = self.getImage(self.ori_path_3)
        ori_slicer3 = (int)(data3.shape[2] / 2)
        data4 = self.getImage(self.ori_path_4)
        ori_slicer4 = (int)(data4.shape[2] / 2)

        upDateCavana(self.F, data1, ori_slicer1)
        upDateCavana(self.F2, data2, ori_slicer2)
        upDateCavana(self.F3, data3, ori_slicer3)
        upDateCavana(self.F4, data4, ori_slicer4)

    def RadiomicsPart(self):
        # radiomics函数调用
        self.nii_path = ' '
        self.RadiomicsFunc = RadiomicsFunc()

        self.dia.radioButton_Para_Cus.clicked.connect(lambda: self.RadiomicsFunc.cusParaSet(self.lineEdit_fileName))

        if self.dia.groupBox_2.isChecked:
            self.RadiomicsFunc.setupParaFile()
        self.dia.comboBox_ImgType.currentIndexChanged.connect(lambda: self.Change(
            self.comboBox_ImgType.currentIndex()))#检查是否为第一项，如果如果是，允许选择特征
        self.dia.comboBox_ImgType.currentIndexChanged.connect(lambda: self.RadiomicsFunc.ImgType_init(
            self.comboBox_ImgType.currentIndex()))  # currentIndex返回当前项的序号(int)，第一个项的序号为0

        self.dia.comboBox_ImgType.currentIndexChanged.connect(lambda: self.stackedWidget.setCurrentIndex(
            self.comboBox_ImgType.currentIndex()-1))#对应换页

        self.PycheckStatus()
        self.save_name = 'Features_Result' #给特征文件一个初始的名字

        self.save_path = ' '
        self.dia.pushButton_4.clicked.connect(
            lambda: self.getSavePath())  # 获取文件输出地址
        self.dia.lineEdit_fileName.textEdited[str].connect(lambda: self.onChange1(self.dia.lineEdit_fileName)) # 实时获取文件名

        try:
            self.dia.pushButton_6.clicked.connect(lambda: self.RadiomicsFunc.outPuts(self.images,
                                                                                 self.save_name,
                                                                                 self.nii_path, False,
                                                                                 self.ori_path,
                                                                                 self.label_path,
                                                                                 self.save_path,
                                                                                self.ori_path_2,self.ori_path_3,self.ori_path_4))  # 不保存数据
            self.dia.pushButton_2.clicked.connect(lambda: self.RadiomicsFunc.outPuts(self.images,
                                                                                 self.save_name,
                                                                                 self.nii_path, True,
                                                                                 self.ori_path,
                                                                                 self.label_path,
                                                                                 self.save_path,
                                                                                self.ori_path_2,self.ori_path_3,self.ori_path_4))  # 保存数据
        except RuntimeError:
            pass

    def DLPart(self):
        self.batch_path = ' '
        self.dia.pushButton_8.clicked.connect(
            lambda: self.getBatchPath())  # 批量导入

        #选择网络类型,序号从0开始
        self.dia.comboBox_NetType.currentIndexChanged.connect(lambda: self.DlpFeatures.netSet(
            self.comboBox_NetType.currentIndex()))

        self.DlpFeatures = DLPExtractor(self.dia.progressBar)

        #选择网络

        # 实时获取文件名
        self.dia.lineEdit_fileName_3.textEdited[str].connect(lambda: self.onChange1(self.dia.lineEdit_fileName_3))

        # 点击Apply图片转化
        self.save_path = './Result/'
        self.dia.pushButton_7.clicked.connect(
            lambda: self.DlpFeatures.converts(self.batch_path,self.save_path,self.dia.textBrowser,self.save_name) ) #nii_path, outputfile
        print(self.batch_path)

    def RadioML(self):
        radioML = RadioML(self.feaNum,self.dia.textBrowser_4,self.dia.label_sample,self.dia.label_OS)
        self.dia.pushButton_5.clicked.connect(lambda: radioML.Features())
        # self.dia.pushButton_11.clicked.connect(lambda: self.dia.label_sample.setPixmap(QPixmap(r"lib\Estimate\Sample_KM.png")))
        self.dia.pushButton_11.clicked.connect(
            lambda: radioML.Score())

        #自训练
        self.dia.lineEdit_features.textEdited[str].connect(
            lambda: self.onChange2(self.dia.lineEdit_features))  # 实时获取文件名
        self.dia.pushButton_12.clicked.connect(
            lambda: radioML.Train())
        self.dia.pushButton_10.clicked.connect(
            lambda: radioML.Features())
        self.dia.pushButton_13.clicked.connect(
            lambda: radioML.Score())

    def PredictionPart(self):

        self.dia.pushButton_upload_radiomics.clicked.connect(
            lambda: getRadiomicsFile()) # 上传radiomics特征
        self.dia.pushButton_output_gene_path.clicked.connect(
            lambda: getSavePath()) # 基因表达量预测结果输出
        self.dia.pushButton_start_predict.clicked.connect(
            lambda: get_input_text_and_run())
        self.dia.pushButton_save_predict.clicked.connect(
            lambda: get_input_text_and_run(flag=True)) # 获取输入文本
        self.showDrugs()


        def getSavePath():
            self.output_gene_path = QFileDialog.getExistingDirectory(self, "choose save path", "./")

        def getRadiomicsFile():
            # self.RFs_file = QFileDialog.getOpenFileName(self, "choose Radiomics File", ".csv")
            options = QFileDialog.Options()
            options |= QFileDialog.ExistingFiles
            self.RFs_file, _ = QFileDialog.getOpenFileNames(self, "QFileDialog.getOpenFileNames()", "./",
                                                    "csv(*.csv)", options=options)

        def get_input_text_and_run(flag=False):
            prediction = RunPredict(self.dia,
                                    self.RFs_file)
            RunPredict.main(prediction)
            # text = self.line_edit.text()

    def showDrugs(self):
        Genes = ['CD40','KYNU','TRAM2','CCRL2','JUNB','FCGR2A','MBOAT1']
        for gen in Genes:
            table_name = f"{gen}_tableWidget"
            tableWidget = getattr(self.dia, table_name)
            tableWidget.horizontalHeader().setHighlightSections(False)
            tableWidget.horizontalHeader().setStyleSheet("QHeaderView::section { background-color: rgb(26, 35, 50); color: rgb(137,195,235); }")

            # 读取Excel文件
            wb = load_workbook(f"Script/Express_Visualize/gene_drugs/drugs_for_{gen}.xlsx")
            sheet = wb.active

            # 获取行数和列数
            rows = sheet.max_row
            cols = sheet.max_column

            # 设置表格的行数和列数
            tableWidget.setRowCount(rows)
            tableWidget.setColumnCount(cols)

            # 读取数据并显示在表格中
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    item = QTableWidgetItem(str(cell_value))
                    tableWidget.setItem(row - 1, col - 1, item)

        self.dia.show()

    def getBatchPath(self): #批量导入
        batch_path = QFileDialog.getExistingDirectory(self, "choose DLP-analysing directory", "./")
        batch_path =  batch_path + '/' #在这里提取了地址之后给他加一个下级，转字符，这样可以直接调用下部文件夹
        self.batch_path =  batch_path
        #print(self.batch_path)

    def getSavePath(self):  # 获取存储地址
        self.save_path = QFileDialog.getExistingDirectory(self, "choose save path", "./")

    def onChange1(self,lineEdit): #实时获取文件名
        VAR = lineEdit.text()# 设置保存特征的csv文件名
        self.save_name = VAR
        #print(self.VAR)

    def onChange2(self,lineEdit): #实时获取RadioML
        VAR = lineEdit.text()  # 设置保存特征的csv文件名
        self.feaNum = VAR

    def Change(self,tag):
        if tag == 1:
            self.groupBox.setCheckable(True)

    def ChangePage(self,item): #list中点击对应右侧换行
        index = self.dia.listWidget.row(item)
        if index==2:
            self.display1()
        elif index==1:
            self.display2()
        #elif index==3:
            #self.display3()
        elif index == 0:
            self.display5()
        index = self.dia.listWidget_2.row(item)
        if index== 0:
            self.display6()
            self.stackedWidget_3.setCurrentIndex(4)
        if index== 1:
            self.display4()

    def display1(self): # Dlp
        self.stackedWidget_2.setCurrentIndex(1)

    def display2(self): #Pyradiomics
        self.stackedWidget_2.setCurrentIndex(2)

    def display3(self):#Here
        self.stackedWidget_2.setCurrentIndex(3)

    def display4(self): #RadioML
        self.stackedWidget_2.setCurrentIndex(4)

    def display6(self): #RadioML2
        self.stackedWidget_2.setCurrentIndex(5)

    def display5(self): # Wel
        self.stackedWidget_2.setCurrentIndex(0)

    def PycheckStatus(self):
        self.dia.checkBox_Nor1.stateChanged.connect(lambda: self.RadiomicsFunc.Add_Nor(self.dia.checkBox_Nor1.isChecked()))
        self.dia.checkBox_Nor2.stateChanged.connect(lambda: self.RadiomicsFunc.Add_Nor(self.dia.checkBox_Nor2.isChecked()))
        self.dia.checkBox_Nor3.stateChanged.connect(lambda: self.RadiomicsFunc.Add_Nor(self.dia.checkBox_Nor3.isChecked()))
        self.dia.checkBox_Nor4.stateChanged.connect(lambda: self.RadiomicsFunc.Add_Nor(self.dia.checkBox_Nor4.isChecked()))
        self.dia.checkBox_15.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_FirstSpe(self.dia.checkBox_15.isChecked()))
        self.dia.checkBox_14.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_FirstSpe2(self.dia.checkBox_14.isChecked()))
        self.dia.checkBox_12.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_NFirstSpe2(self.dia.checkBox_12.isChecked()))
        self.dia.checkBox_13.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_FirstSpe2(self.dia.checkBox_13.isChecked()))
        self.dia.checkBox_Mas4.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Mask(self.dia.checkBox_Mas4.isChecked()))
        self.dia.checkBox_Mas3.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Mask(self.dia.checkBox_Mas3.isChecked()))
        self.dia.checkBox_Mas2.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Mask(self.dia.checkBox_Mas2.isChecked()))
        self.dia.checkBox_Mas1.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Mask(self.dia.checkBox_Mas1.isChecked()))
        self.dia.checkBox_Mis2.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Misc(self.dia.checkBox_Mis2.isChecked()))
        self.dia.checkBox_Res1.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Resampling(self.dia.checkBox_Res1.isChecked()))
        self.dia.checkBox_Res3.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Resampling(self.dia.checkBox_Res3.isChecked()))
        self.dia.checkBox_2D1.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_2D(self.dia.checkBox_2D.isChecked()))
        self.dia.checkBox_Voxel.stateChanged.connect(
            lambda: self.RadiomicsFunc.Add_Voxel(self.dia.checkBox_Voxel.isChecked()))

    def bindButton(self):
        file_name = QFileDialog.getOpenFileName(None, "Open File", "./", "nii(*.nii.gz;*.nii;*.nrrd)")
        self.nii_path = file_name[0]  # 提取文件路径
        try:
            self.data_mask = nib.load(Path(self.nii_path))
        except nib.filebasedimages.ImageFileError:
            pass
        else:
            self.data2 = self.data_mask.get_fdata()
            SimpleView.VisNii(self, self.nii_path, self.ori_path)

    def bindSlider(self):
        slice_idx = self.horizontalSlider.value()
        self.showimage(slice_idx)

    def quit_button(self):
        quit()

    def mousePressEvent(self, event):  # 鼠标左键按下时获取鼠标坐标,按下右键取消
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()
            event.accept()
        elif event.button() == Qt.RightButton:
            self.m_flag = False

    def mouseMoveEvent(self, QMouseEvent):  # 鼠标在按下左键的情况下移动时,根据坐标移动界面
        try:
            if Qt.LeftButton and self.m_flag:
                self.move(QMouseEvent.globalPos() - self.m_Position)
                QMouseEvent.accept()
        except AttributeError:
            pass

    def mouseReleaseEvent(self, QMouseEvent):  # 鼠标按键释放时,取消移动
        self.m_flag = False


